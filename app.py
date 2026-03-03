"""
Arcstone Advisory - Executive Financial Intelligence System
Tier I Sample Dashboard - MBB Methodology
"""
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

st.set_page_config(page_title="Arcstone Advisory - Financial Intelligence", page_icon="◆", layout="wide", initial_sidebar_state="expanded")

NAVY="#1B2A4A"; ACCENT="#4A6FA5"; CHARCOAL="#3A3F47"; SLATE="#6B7280"
LIGHT="#F8F9FB"; BORDER="#E5E7EB"; GREEN="#059669"; RED="#DC2626"; AMBER="#D97706"

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
html, body, [class*="st-"]{font-family:'Inter',-apple-system,BlinkMacSystemFont,sans-serif}
.main .block-container{padding:2rem 3rem;max-width:1400px}
section[data-testid="stSidebar"]>div{padding-top:1.5rem}
h1,h2,h3{color:#1B2A4A !important;font-weight:700 !important}
h1{font-size:26px !important;letter-spacing:-0.5px !important}
h2{font-size:18px !important;margin-top:1.5rem !important}
h3{font-size:15px !important}
[data-testid="stMetric"]{background:#F8F9FB;border:1px solid #E5E7EB;border-radius:8px;padding:20px 24px}
[data-testid="stMetricLabel"]{font-size:10px !important;font-weight:600 !important;letter-spacing:1.5px !important;text-transform:uppercase !important;color:#6B7280 !important}
[data-testid="stMetricValue"]{font-size:26px !important;font-weight:700 !important;color:#1B2A4A !important}
[data-testid="stMetricDelta"]>div{font-size:12px !important}
[data-testid="stSidebar"]{background:#1B2A4A}
[data-testid="stSidebar"] *{color:rgba(255,255,255,0.85) !important}
[data-testid="stSidebar"] .stSelectbox label,[data-testid="stSidebar"] .stRadio label{font-size:10px !important;font-weight:600 !important;letter-spacing:1.5px !important;text-transform:uppercase !important}
[data-testid="stSidebar"] hr{border-color:rgba(255,255,255,0.12) !important}
.tag{font-size:10px;font-weight:600;letter-spacing:2.5px;text-transform:uppercase;color:#4A6FA5}
.divider{border:none;border-top:1px solid #E5E7EB;margin:1.5rem 0}
.insight{background:#F8F9FB;border-left:3px solid #4A6FA5;padding:16px 20px;border-radius:0 6px 6px 0;margin:12px 0;font-size:14px;color:#3A3F47;line-height:1.6}
#MainMenu,footer,header{visibility:hidden}
</style>
""", unsafe_allow_html=True)

# ━━━ DATA ━━━
@st.cache_data
def load_data():
    np.random.seed(42)
    months=pd.date_range("2024-01-01",periods=12,freq="MS")
    rev=np.array([420,445,460,485,470,510,530,520,555,540,570,590])*1000
    cg=np.array([243600,253650,271400,271600,272600,290700,296800,301600,305250,307800,319200,324500])
    gp=rev-cg
    pers=np.array([52,52,53,53,53,54,54,54,55,55,55,56])*1000
    rnt=np.full(12,12000); mkt=np.array([8,9,7.5,10,8.5,9.5,11,9,10.5,9,10,11.5])*1000
    oox=np.array([15,14.5,16,15.5,14,16.5,15,15.5,16,14.5,15.5,16])*1000
    topx=pers+rnt+mkt+oox; ebitda=gp-topx; da=np.full(12,8000); ebit=ebitda-da
    intr=np.linspace(3500,2400,12).astype(int); ebt=ebit-intr
    tx=np.maximum(0,(ebt*0.25)).astype(int); ni=ebt-tx

    cash=np.array([180,195,188,210,205,225,240,235,255,248,265,285])*1000
    rec=np.array([125,132,138,145,140,152,158,155,166,161,170,176])*1000
    inv=np.array([95,98,102,100,105,108,110,106,112,115,118,120])*1000
    oca=np.full(12,25000); tca=cash+rec+inv+oca
    pay=np.array([85,88,92,90,95,98,100,96,102,105,108,110])*1000
    sd=np.linspace(50000,28000,12).astype(int); ocl=np.full(12,20000); tcl=pay+sd+ocl
    ld=np.linspace(200000,156000,12).astype(int)
    ta=tca+380000; tl=tcl+ld+10000; eq=ta-tl

    facs=[.82,.85,.79,.91,.83,.88,.86,.80,.93,.84,.87,.90]
    opcf=np.array([int(e*f) for e,f in zip(ebitda,facs)])
    icf=np.array([-28000,-35000,-22000,-41000,-19000,-37000,-30000,-25000,-42000,-20000,-33000,-38000])
    fcf=np.array([-15000,-12000,-18000,-10000,-20000,-14000,-16000,-22000,-11000,-19000,-13000,-17000])
    ncf=opcf+icf+fcf
    brev=rev*np.array([1.02,0.98,1.03,0.97,1.01,1.04,0.99,1.02,0.98,1.03,1.01,0.97])
    bebitda=ebitda*np.array([1.05,0.96,1.02,0.98,1.04,1.01,0.97,1.03,0.99,1.02,1.00,0.98])

    return pd.DataFrame({
        "Month":months,"Label":[m.strftime("%b %y") for m in months],
        "Revenue":rev,"COGS":cg,"GP":gp,"Personnel":pers,"Rent":rnt,"Marketing":mkt,
        "Other_OpEx":oox,"TotalOpEx":topx,"EBITDA":ebitda,"DA":da,"EBIT":ebit,
        "Interest":intr,"EBT":ebt,"Tax":tx,"NI":ni,
        "Cash":cash,"Rec":rec,"Inv":inv,"TCA":tca,"Pay":pay,"SD":sd,"TCL":tcl,"LD":ld,
        "TA":ta,"TL":tl,"Eq":eq,"OpCF":opcf,"InvCF":icf,"FinCF":fcf,"NetCF":ncf,
        "GM":gp/rev*100,"EM":ebitda/rev*100,"NM":ni/rev*100,
        "CR":tca/tcl,"QR":(tca-inv)/tcl,"DE":tl/eq,"ROE":ni/eq*100,"ROA":ni/ta*100,
        "DSO":rec/rev*30,"DIO":inv/cg*30,"DPO":pay/cg*30,
        "CCC":rec/rev*30+inv/cg*30-pay/cg*30,
        "BRev":brev,"BEBITDA":bebitda,
    })

df=load_data()

def mbb(fig,title="",h=380,yp="€",ys="",yf=","):
    fig.update_layout(
        title=dict(text=f"<b>{title}</b>",font=dict(size=14,color=NAVY,family="Inter"),x=0,xanchor="left"),
        font=dict(family="Inter",size=11,color=SLATE),plot_bgcolor="white",paper_bgcolor="white",
        margin=dict(l=50,r=20,t=50,b=40),
        legend=dict(orientation="h",yanchor="bottom",y=1.02,xanchor="right",x=1,font=dict(size=10,color=SLATE),bgcolor="rgba(0,0,0,0)"),
        xaxis=dict(gridcolor="#F3F4F6",linecolor=BORDER,tickfont=dict(size=10)),
        yaxis=dict(gridcolor="#F3F4F6",linecolor=BORDER,tickprefix=yp,ticksuffix=ys,tickformat=yf,tickfont=dict(size=10)),
        hovermode="x unified",height=h,
        hoverlabel=dict(bgcolor="white",font_size=11,font_family="Inter",bordercolor=BORDER))
    return fig

def pct_d(c,p): return f"{((c-p)/abs(p)*100):+.1f}%" if p!=0 else "–"
def pp_d(c,p): return f"{c-p:+.1f}pp"

# ━━━ SIDEBAR ━━━
with st.sidebar:
    st.markdown("#### ◆ ARCSTONE ADVISORY")
    st.markdown("*Financial Intelligence System*")
    st.markdown("---")
    period=st.selectbox("REPORTING PERIOD",["Full Year 2024","H1 2024","H2 2024","Q1 2024","Q2 2024","Q3 2024","Q4 2024"])
    view=st.radio("ANALYSIS VIEW",["Executive Summary","P&L Analysis","Margin Deep Dive","Balance Sheet","Cash Flow & Liquidity","Working Capital","Variance Analysis"])
    st.markdown("---")
    st.markdown(f"**Last Updated:** {datetime.now().strftime('%d %b %Y')}")
    st.markdown("**Data Source:** ERP + Accounting")
    st.markdown("**Currency:** EUR")
    st.markdown("---")
    st.caption("Sample data for demonstration.\nEach client system is fully customized.")

pm={"Full Year 2024":slice(None),"H1 2024":slice(0,6),"H2 2024":slice(6,12),"Q1 2024":slice(0,3),"Q2 2024":slice(3,6),"Q3 2024":slice(6,9),"Q4 2024":slice(9,12)}
d=df.iloc[pm[period]].reset_index(drop=True); n=len(d)
L=d.iloc[-1]; P=d.iloc[-2] if n>1 else L

st.markdown('<p class="tag">TIER I — EXECUTIVE FINANCIAL SYSTEM</p>',unsafe_allow_html=True)
st.markdown(f"# Financial Intelligence — {period}")
st.markdown('<hr class="divider">',unsafe_allow_html=True)

# ══════ EXECUTIVE SUMMARY ══════
if view=="Executive Summary":
    c1,c2,c3,c4,c5,c6=st.columns(6)
    c1.metric("Revenue",f"€{L['Revenue']/1000:.0f}K",pct_d(L['Revenue'],P['Revenue']))
    c2.metric("Gross Margin",f"{L['GM']:.1f}%",pp_d(L['GM'],P['GM']))
    c3.metric("EBITDA",f"€{L['EBITDA']/1000:.0f}K",pct_d(L['EBITDA'],P['EBITDA']))
    c4.metric("EBITDA Margin",f"{L['EM']:.1f}%",pp_d(L['EM'],P['EM']))
    c5.metric("Net Income",f"€{L['NI']/1000:.0f}K",pct_d(L['NI'],P['NI']))
    c6.metric("Cash",f"€{L['Cash']/1000:.0f}K",pct_d(L['Cash'],P['Cash']))
    st.markdown('<hr class="divider">',unsafe_allow_html=True)

    rg=(d['Revenue'].iloc[-1]-d['Revenue'].iloc[0])/d['Revenue'].iloc[0]*100
    st.markdown(f'<div class="insight"><strong>Executive Insight:</strong> Revenue grew <strong>{rg:.1f}%</strong> over the period. Average EBITDA margin: <strong>{d["EM"].mean():.1f}%</strong>. Cash strengthened by <strong>€{(d["Cash"].iloc[-1]-d["Cash"].iloc[0])/1000:.0f}K</strong>. Cash conversion cycle: <strong>{L["CCC"]:.0f} days</strong>.</div>',unsafe_allow_html=True)

    c1,c2=st.columns(2)
    with c1:
        fig=go.Figure()
        fig.add_trace(go.Bar(x=d["Label"],y=d["Revenue"]/1000,name="Revenue",marker_color=NAVY,opacity=0.9))
        fig.add_trace(go.Scatter(x=d["Label"],y=d["EBITDA"]/1000,name="EBITDA",line=dict(color=ACCENT,width=2.5),mode="lines+markers",marker=dict(size=5)))
        fig=mbb(fig,"Revenue & EBITDA (€K)");fig.update_yaxes(ticksuffix="K")
        st.plotly_chart(fig,use_container_width=True)
    with c2:
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=d["Label"],y=d["GM"],name="Gross",line=dict(color=NAVY,width=2.5),mode="lines+markers",marker=dict(size=5)))
        fig.add_trace(go.Scatter(x=d["Label"],y=d["EM"],name="EBITDA",line=dict(color=ACCENT,width=2.5),mode="lines+markers",marker=dict(size=5)))
        fig.add_trace(go.Scatter(x=d["Label"],y=d["NM"],name="Net",line=dict(color="#94A3B8",width=2),mode="lines+markers",marker=dict(size=4)))
        fig=mbb(fig,"Margin Evolution (%)");fig.update_yaxes(tickprefix="",ticksuffix="%",tickformat=".1f")
        st.plotly_chart(fig,use_container_width=True)

    c3,c4=st.columns(2)
    with c3:
        fig=go.Figure(go.Waterfall(
            x=["Revenue","COGS","Gross Profit","OpEx","EBITDA","D&A","Interest","Tax","Net Income"],
            y=[L['Revenue'],-L['COGS'],0,-L['TotalOpEx'],0,-L['DA'],-L['Interest'],-L['Tax'],0],
            measure=["absolute","relative","total","relative","total","relative","relative","relative","total"],
            connector=dict(line=dict(color=BORDER,width=1)),
            increasing=dict(marker=dict(color=NAVY)),decreasing=dict(marker=dict(color=RED)),totals=dict(marker=dict(color=ACCENT)),
            textposition="outside",
            text=[f"€{v/1000:.0f}K" for v in [L['Revenue'],L['COGS'],L['GP'],L['TotalOpEx'],L['EBITDA'],L['DA'],L['Interest'],L['Tax'],L['NI']]],
            textfont=dict(size=9,color=SLATE)))
        fig=mbb(fig,f"P&L Bridge — {L['Label']}",h=400);fig.update_yaxes(visible=False);fig.update_xaxes(tickfont=dict(size=9))
        st.plotly_chart(fig,use_container_width=True)
    with c4:
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=d["Label"],y=d["Cash"]/1000,fill="tozeroy",fillcolor="rgba(74,111,165,0.08)",line=dict(color=ACCENT,width=2.5),name="Cash",mode="lines+markers",marker=dict(size=5)))
        fig=mbb(fig,"Cash Position (€K)");fig.update_yaxes(ticksuffix="K")
        st.plotly_chart(fig,use_container_width=True)

    st.markdown("### Period Summary")
    sm=pd.DataFrame({"Metric":["Total Revenue","Total Gross Profit","Total EBITDA","Total Net Income","Avg Gross Margin","Avg EBITDA Margin","Avg Net Margin","Closing Cash"],
        "Value":[f"€{d['Revenue'].sum()/1000:,.0f}K",f"€{d['GP'].sum()/1000:,.0f}K",f"€{d['EBITDA'].sum()/1000:,.0f}K",f"€{d['NI'].sum()/1000:,.0f}K",f"{d['GM'].mean():.1f}%",f"{d['EM'].mean():.1f}%",f"{d['NM'].mean():.1f}%",f"€{d['Cash'].iloc[-1]/1000:,.0f}K"]})
    st.dataframe(sm,use_container_width=True,hide_index=True)

# ══════ P&L ANALYSIS ══════
elif view=="P&L Analysis":
    c1,c2,c3,c4=st.columns(4)
    c1.metric("Total Revenue",f"€{d['Revenue'].sum()/1000:,.0f}K")
    c2.metric("Total Gross Profit",f"€{d['GP'].sum()/1000:,.0f}K")
    c3.metric("Total EBITDA",f"€{d['EBITDA'].sum()/1000:,.0f}K")
    c4.metric("Total Net Income",f"€{d['NI'].sum()/1000:,.0f}K")
    st.markdown('<hr class="divider">',unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        fig=go.Figure()
        fig.add_trace(go.Bar(x=d["Label"],y=d["Revenue"]/1000,name="Revenue",marker_color=NAVY,opacity=0.9))
        fig.add_trace(go.Bar(x=d["Label"],y=-d["COGS"]/1000,name="COGS",marker_color=RED,opacity=0.7))
        fig.add_trace(go.Bar(x=d["Label"],y=-d["TotalOpEx"]/1000,name="OpEx",marker_color=AMBER,opacity=0.7))
        fig=mbb(fig,"Revenue vs Cost Structure (€K)");fig.update_layout(barmode="relative");fig.update_yaxes(ticksuffix="K")
        st.plotly_chart(fig,use_container_width=True)
    with c2:
        fig=go.Figure()
        fig.add_trace(go.Bar(x=d["Label"],y=d["Personnel"]/1000,name="Personnel",marker_color=NAVY))
        fig.add_trace(go.Bar(x=d["Label"],y=d["Rent"]/1000,name="Rent",marker_color=ACCENT))
        fig.add_trace(go.Bar(x=d["Label"],y=d["Marketing"]/1000,name="Marketing",marker_color="#94A3B8"))
        fig.add_trace(go.Bar(x=d["Label"],y=d["Other_OpEx"]/1000,name="Other",marker_color="#CBD5E1"))
        fig=mbb(fig,"OpEx Breakdown (€K)");fig.update_layout(barmode="stack");fig.update_yaxes(ticksuffix="K")
        st.plotly_chart(fig,use_container_width=True)
    st.markdown("### Profit & Loss Statement")
    pl=pd.DataFrame({"Month":d["Label"],"Revenue":d["Revenue"].apply(lambda x:f"€{x/1000:,.0f}K"),"COGS":d["COGS"].apply(lambda x:f"(€{x/1000:,.0f}K)"),"Gross Profit":d["GP"].apply(lambda x:f"€{x/1000:,.0f}K"),"GM%":d["GM"].apply(lambda x:f"{x:.1f}%"),"OpEx":d["TotalOpEx"].apply(lambda x:f"(€{x/1000:,.0f}K)"),"EBITDA":d["EBITDA"].apply(lambda x:f"€{x/1000:,.0f}K"),"EM%":d["EM"].apply(lambda x:f"{x:.1f}%"),"Net Income":d["NI"].apply(lambda x:f"€{x/1000:,.0f}K"),"NM%":d["NM"].apply(lambda x:f"{x:.1f}%")})
    st.dataframe(pl,use_container_width=True,hide_index=True)

# ══════ MARGIN DEEP DIVE ══════
elif view=="Margin Deep Dive":
    c1,c2,c3=st.columns(3)
    c1.metric("Avg Gross Margin",f"{d['GM'].mean():.1f}%",f"σ = {d['GM'].std():.2f}pp")
    c2.metric("Avg EBITDA Margin",f"{d['EM'].mean():.1f}%",f"σ = {d['EM'].std():.2f}pp")
    c3.metric("Avg Net Margin",f"{d['NM'].mean():.1f}%",f"σ = {d['NM'].std():.2f}pp")
    st.markdown('<hr class="divider">',unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        fig=make_subplots(specs=[[{"secondary_y":True}]])
        fig.add_trace(go.Bar(x=d["Label"],y=d["Revenue"]/1000,name="Revenue",marker_color=NAVY,opacity=0.3),secondary_y=False)
        fig.add_trace(go.Scatter(x=d["Label"],y=d["GM"],name="Gross Margin",line=dict(color=NAVY,width=2.5),mode="lines+markers",marker=dict(size=5)),secondary_y=True)
        fig.add_trace(go.Scatter(x=d["Label"],y=d["EM"],name="EBITDA Margin",line=dict(color=ACCENT,width=2.5),mode="lines+markers",marker=dict(size=5)),secondary_y=True)
        fig=mbb(fig,"Revenue Scale vs Margins");fig.update_yaxes(tickprefix="€",ticksuffix="K",secondary_y=False);fig.update_yaxes(ticksuffix="%",tickformat=".1f",secondary_y=True,gridcolor="rgba(0,0,0,0)")
        st.plotly_chart(fig,use_container_width=True)
    with c2:
        cp=d['COGS']/d['Revenue']*100;op=d['TotalOpEx']/d['Revenue']*100
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=d["Label"],y=cp,name="COGS %",stackgroup="one",fillcolor="rgba(27,42,74,0.15)",line=dict(color=NAVY,width=0)))
        fig.add_trace(go.Scatter(x=d["Label"],y=op,name="OpEx %",stackgroup="one",fillcolor="rgba(74,111,165,0.15)",line=dict(color=ACCENT,width=0)))
        fig=mbb(fig,"Cost as % of Revenue");fig.update_yaxes(tickprefix="",ticksuffix="%")
        st.plotly_chart(fig,use_container_width=True)
    st.markdown(f'<div class="insight"><strong>Margin Volatility:</strong> Gross Margin σ = <strong>{d["GM"].std():.2f}pp</strong> | EBITDA Margin σ = <strong>{d["EM"].std():.2f}pp</strong> | Net Margin σ = <strong>{d["NM"].std():.2f}pp</strong>. Target: σ < 2.0pp.</div>',unsafe_allow_html=True)

# ══════ BALANCE SHEET ══════
elif view=="Balance Sheet":
    c1,c2,c3,c4=st.columns(4)
    c1.metric("Total Assets",f"€{L['TA']/1000:,.0f}K",pct_d(L['TA'],P['TA']))
    c2.metric("Current Ratio",f"{L['CR']:.2f}x",f"{L['CR']-P['CR']:+.2f}")
    c3.metric("Quick Ratio",f"{L['QR']:.2f}x",f"{L['QR']-P['QR']:+.2f}")
    c4.metric("Debt/Equity",f"{L['DE']:.2f}x",f"{L['DE']-P['DE']:+.2f}",delta_color="inverse")
    st.markdown('<hr class="divider">',unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        fig=go.Figure()
        fig.add_trace(go.Bar(x=d["Label"],y=d["Cash"]/1000,name="Cash",marker_color=GREEN))
        fig.add_trace(go.Bar(x=d["Label"],y=d["Rec"]/1000,name="Receivables",marker_color=NAVY))
        fig.add_trace(go.Bar(x=d["Label"],y=d["Inv"]/1000,name="Inventory",marker_color=ACCENT))
        fig=mbb(fig,"Current Asset Composition (€K)");fig.update_layout(barmode="stack");fig.update_yaxes(ticksuffix="K")
        st.plotly_chart(fig,use_container_width=True)
    with c2:
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=d["Label"],y=d["CR"],name="Current Ratio",line=dict(color=NAVY,width=2.5),mode="lines+markers",marker=dict(size=5)))
        fig.add_trace(go.Scatter(x=d["Label"],y=d["QR"],name="Quick Ratio",line=dict(color=ACCENT,width=2.5),mode="lines+markers",marker=dict(size=5)))
        fig.add_hline(y=1.5,line=dict(color=GREEN,dash="dash",width=1),annotation_text="Target: 1.5x")
        fig.add_hline(y=1.0,line=dict(color=RED,dash="dash",width=1),annotation_text="Min: 1.0x")
        fig=mbb(fig,"Liquidity Ratios");fig.update_yaxes(tickprefix="",ticksuffix="x",tickformat=".2f")
        st.plotly_chart(fig,use_container_width=True)

# ══════ CASH FLOW ══════
elif view=="Cash Flow & Liquidity":
    c1,c2,c3,c4=st.columns(4)
    c1.metric("Operating CF",f"€{L['OpCF']/1000:.0f}K",pct_d(L['OpCF'],P['OpCF']))
    c2.metric("Net Cash Flow",f"€{L['NetCF']/1000:.0f}K")
    c3.metric("Cash Position",f"€{L['Cash']/1000:.0f}K",pct_d(L['Cash'],P['Cash']))
    c4.metric("CF / Revenue",f"{L['OpCF']/L['Revenue']*100:.1f}%")
    st.markdown('<hr class="divider">',unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        fig=go.Figure()
        fig.add_trace(go.Bar(x=d["Label"],y=d["OpCF"]/1000,name="Operating",marker_color=NAVY))
        fig.add_trace(go.Bar(x=d["Label"],y=d["InvCF"]/1000,name="Investing",marker_color=RED,opacity=0.7))
        fig.add_trace(go.Bar(x=d["Label"],y=d["FinCF"]/1000,name="Financing",marker_color="#94A3B8"))
        fig.add_trace(go.Scatter(x=d["Label"],y=d["NetCF"]/1000,name="Net CF",line=dict(color=ACCENT,width=2.5,dash="dot"),mode="lines+markers",marker=dict(size=6)))
        fig=mbb(fig,"Cash Flow Components (€K)");fig.update_layout(barmode="group");fig.update_yaxes(ticksuffix="K")
        st.plotly_chart(fig,use_container_width=True)
    with c2:
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=d["Label"],y=d["Cash"]/1000,fill="tozeroy",fillcolor="rgba(5,150,105,0.08)",line=dict(color=GREEN,width=2.5),name="Cash",mode="lines+markers",marker=dict(size=5)))
        fig=mbb(fig,"Cash Position (€K)");fig.update_yaxes(ticksuffix="K")
        st.plotly_chart(fig,use_container_width=True)
    st.markdown(f'<div class="insight"><strong>Cash Conversion:</strong> Operating CF = <strong>{d["OpCF"].sum()/d["EBITDA"].sum()*100:.0f}%</strong> of EBITDA. Cumulative net CF: <strong>€{d["NetCF"].sum()/1000:,.0f}K</strong>.</div>',unsafe_allow_html=True)

# ══════ WORKING CAPITAL ══════
elif view=="Working Capital":
    c1,c2,c3,c4=st.columns(4)
    c1.metric("DSO",f"{L['DSO']:.0f} days",f"{L['DSO']-P['DSO']:+.0f}d",delta_color="inverse")
    c2.metric("DIO",f"{L['DIO']:.0f} days",f"{L['DIO']-P['DIO']:+.0f}d",delta_color="inverse")
    c3.metric("DPO",f"{L['DPO']:.0f} days",f"{L['DPO']-P['DPO']:+.0f}d")
    c4.metric("CCC",f"{L['CCC']:.0f} days",f"{L['CCC']-P['CCC']:+.0f}d",delta_color="inverse")
    st.markdown('<hr class="divider">',unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=d["Label"],y=d["DSO"],name="DSO",line=dict(color=NAVY,width=2.5),mode="lines+markers",marker=dict(size=5)))
        fig.add_trace(go.Scatter(x=d["Label"],y=d["DIO"],name="DIO",line=dict(color=ACCENT,width=2.5),mode="lines+markers",marker=dict(size=5)))
        fig.add_trace(go.Scatter(x=d["Label"],y=d["DPO"],name="DPO",line=dict(color=GREEN,width=2.5),mode="lines+markers",marker=dict(size=5)))
        fig=mbb(fig,"Working Capital Cycle (Days)");fig.update_yaxes(tickprefix="",ticksuffix="d",tickformat=".0f")
        st.plotly_chart(fig,use_container_width=True)
    with c2:
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=d["Label"],y=d["CCC"],fill="tozeroy",fillcolor="rgba(27,42,74,0.08)",line=dict(color=NAVY,width=2.5),name="CCC",mode="lines+markers",marker=dict(size=5)))
        fig.add_hline(y=d["CCC"].mean(),line=dict(color=ACCENT,dash="dash",width=1),annotation_text=f"Avg: {d['CCC'].mean():.0f}d")
        fig=mbb(fig,"Cash Conversion Cycle (Days)");fig.update_yaxes(tickprefix="",ticksuffix="d",tickformat=".0f")
        st.plotly_chart(fig,use_container_width=True)
    nwc=d["Rec"]+d["Inv"]-d["Pay"]
    st.markdown("### Net Working Capital")
    fig=go.Figure()
    fig.add_trace(go.Bar(x=d["Label"],y=d["Rec"]/1000,name="Receivables",marker_color=NAVY))
    fig.add_trace(go.Bar(x=d["Label"],y=d["Inv"]/1000,name="Inventory",marker_color=ACCENT))
    fig.add_trace(go.Bar(x=d["Label"],y=-d["Pay"]/1000,name="Payables",marker_color=GREEN))
    fig.add_trace(go.Scatter(x=d["Label"],y=nwc/1000,name="Net WC",line=dict(color=RED,width=2.5,dash="dot"),mode="lines+markers",marker=dict(size=6)))
    fig=mbb(fig,"Working Capital Components (€K)",h=350);fig.update_layout(barmode="relative");fig.update_yaxes(ticksuffix="K")
    st.plotly_chart(fig,use_container_width=True)

# ══════ VARIANCE ANALYSIS ══════
elif view=="Variance Analysis":
    rv=(d['Revenue'].sum()-d['BRev'].sum())/d['BRev'].sum()*100
    ev=(d['EBITDA'].sum()-d['BEBITDA'].sum())/d['BEBITDA'].sum()*100
    c1,c2,c3,c4=st.columns(4)
    c1.metric("Revenue vs Budget",f"{rv:+.1f}%","Above" if rv>0 else "Below")
    c2.metric("EBITDA vs Budget",f"{ev:+.1f}%","Above" if ev>0 else "Below")
    c3.metric("Best Month",f"{d.loc[d['Revenue'].idxmax(),'Label']}")
    c4.metric("Worst Month",f"{d.loc[d['Revenue'].idxmin(),'Label']}")
    st.markdown('<hr class="divider">',unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        fig=go.Figure()
        fig.add_trace(go.Bar(x=d["Label"],y=d["Revenue"]/1000,name="Actual",marker_color=NAVY,opacity=0.9))
        fig.add_trace(go.Scatter(x=d["Label"],y=d["BRev"]/1000,name="Budget",line=dict(color=RED,width=2,dash="dash"),mode="lines+markers",marker=dict(size=5)))
        fig=mbb(fig,"Revenue: Actual vs Budget (€K)");fig.update_yaxes(ticksuffix="K")
        st.plotly_chart(fig,use_container_width=True)
    with c2:
        fig=go.Figure()
        fig.add_trace(go.Bar(x=d["Label"],y=d["EBITDA"]/1000,name="Actual",marker_color=NAVY,opacity=0.9))
        fig.add_trace(go.Scatter(x=d["Label"],y=d["BEBITDA"]/1000,name="Budget",line=dict(color=RED,width=2,dash="dash"),mode="lines+markers",marker=dict(size=5)))
        fig=mbb(fig,"EBITDA: Actual vs Budget (€K)");fig.update_yaxes(ticksuffix="K")
        st.plotly_chart(fig,use_container_width=True)
    st.markdown("### Monthly Variance")
    vdf=pd.DataFrame({"Month":d["Label"],"Rev Actual":d["Revenue"].apply(lambda x:f"€{x/1000:,.0f}K"),"Rev Budget":d["BRev"].apply(lambda x:f"€{x/1000:,.0f}K"),"Rev Δ%":((d["Revenue"]-d["BRev"])/d["BRev"]*100).apply(lambda x:f"{x:+.1f}%"),"EBITDA Actual":d["EBITDA"].apply(lambda x:f"€{x/1000:,.0f}K"),"EBITDA Budget":d["BEBITDA"].apply(lambda x:f"€{x/1000:,.0f}K"),"EBITDA Δ%":((d["EBITDA"]-d["BEBITDA"])/d["BEBITDA"]*100).apply(lambda x:f"{x:+.1f}%")})
    st.dataframe(vdf,use_container_width=True,hide_index=True)

# ━━━ DOWNLOAD ━━━
st.markdown('<hr class="divider">',unsafe_allow_html=True)

def make_excel(data):
    buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine='openpyxl') as w:
        pl2=data[["Label","Revenue","COGS","GP","TotalOpEx","EBITDA","NI","GM","EM","NM"]].copy()
        pl2.columns=["Month","Revenue","COGS","Gross Profit","OpEx","EBITDA","Net Income","Gross Margin %","EBITDA Margin %","Net Margin %"]
        pl2.to_excel(w,sheet_name="P&L",index=False)
        bs2=data[["Label","Cash","Rec","Inv","TCA","Pay","SD","TCL","LD","TA","TL","Eq"]].copy()
        bs2.columns=["Month","Cash","Receivables","Inventory","Total CA","Payables","Short Debt","Total CL","Long Debt","Total Assets","Total Liab","Equity"]
        bs2.to_excel(w,sheet_name="Balance Sheet",index=False)
        cf2=data[["Label","OpCF","InvCF","FinCF","NetCF","Cash"]].copy()
        cf2.columns=["Month","Operating CF","Investing CF","Financing CF","Net CF","Cash Position"]
        cf2.to_excel(w,sheet_name="Cash Flow",index=False)
        rt2=data[["Label","GM","EM","NM","CR","QR","DE","ROE","ROA","DSO","DIO","DPO","CCC"]].copy()
        rt2.columns=["Month","Gross Margin %","EBITDA Margin %","Net Margin %","Current Ratio","Quick Ratio","Debt/Equity","ROE %","ROA %","DSO","DIO","DPO","CCC"]
        rt2.to_excel(w,sheet_name="Ratios",index=False)
        for sn in w.sheets:
            ws2=w.sheets[sn]
            for cell in ws2[1]:
                cell.font=Font(name='Calibri',size=10,bold=True,color='FFFFFF')
                cell.fill=PatternFill(start_color='1B2A4A',end_color='1B2A4A',fill_type='solid')
                cell.alignment=Alignment(horizontal='center')
            for col in ws2.columns: ws2.column_dimensions[get_column_letter(col[0].column)].width=16
            for row in ws2.iter_rows(min_row=2):
                for cell in row:
                    cell.font=Font(name='Calibri',size=10)
                    if isinstance(cell.value,(int,float)) and abs(cell.value)>100: cell.number_format='#,##0'
    buf.seek(0); return buf

c1,c2,_=st.columns([1,1,4])
with c1:
    st.download_button("📥 Export Report",make_excel(d),file_name=f"arcstone_financial_{period.replace(' ','_').lower()}.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
with c2:
    tpl=os.path.join(os.path.dirname(os.path.abspath(__file__)),"arcstone_financial_template.xlsx")
    if os.path.exists(tpl):
        with open(tpl,"rb") as f:
            st.download_button("📋 Client Template",f.read(),file_name="arcstone_financial_template.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)

st.markdown("---")
st.caption("◆ Arcstone Advisory — Executive Financial Intelligence System | Sample data | arcstoneadvisory.com")
